import openpyxl
from openpyxl.styles import PatternFill

def getRow(file, sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.max_row


def getColumn(file, sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.max_column


def getData (file, sheetName, rowNum,columnNum):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    return sheet.cell(rowNum, columnNum).value

def writeData (file, sheetName, rowNum, columnNum, data):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    sheet.cell(rowNum, columnNum).value = data
    workBook.save(file)


def setGreenColor(file, sheetName, rowNum, columnNum):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    greenFill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rowNum, columnNum).fill = greenFill
    workBook.save(file)


def setRedColor(file, sheetName, rowNum, columnNum):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook[sheetName]
    redFill = PatternFill(start_color='#FF0000', end_color='#FF0000', fill_type='solid')
    sheet.cell(rowNum, columnNum).fill = redFill
    workBook.save(file)

