import openpyxl

#  Structure of Xcell File
#  File --> Workbook---> Sheets--> Rows-->Cells

file = 'C:\\Users\\Admin\\Desktop\\TestBook.xlsx'
workbook = openpyxl.load_workbook(file)
sheet = workbook['Лист1']
rows = sheet.max_row
columns = sheet.max_column

for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r, c).value, end=' ')
    print()

